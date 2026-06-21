# -*- coding: utf-8 -*-
"""YAML 原稿 → Excel(.xlsx) 生成器（「表が主役」のドキュメント用）。

`docs/<project>/src/*.xlsx.yaml` を正典とし、openpyxl で業務 Excel 様式（ヘッダ塗り・ゼブラ・
細罫線・ウィンドウ枠固定・オートフィルタ・A4 横の印刷設定）を組む。手順書・設計書のような流れる
文書は従来どおり `md2docx.py`（Word）が担当し、本書は画面項目定義 / 入出力定義 / DB 定義 /
テスト仕様といった**表が主役の文書**を受け持つ。

フォント規約（CLAUDE.md「Word ドキュメント」）と共通: 本文 `BIZ UDPGothic`、等幅 `BIZ UDGothic`。
配色も Word 版デザインシステムと共通トークン（ACCENT=#1F5C99 等）。結果色（合格=緑）のみ Excel 用に追加。

依存: openpyxl のみ。**PyYAML はオフラインバンドルに無いため使わず**、必要な YAML サブセットだけを
自前パーサ（`parse_yaml`）で読む（`md2docx.py` の「外部ライブラリを足さない」方針と整合）。

入力フォーマット（YAML サブセット・1 ファイル 1 ブック）:
  out:   出力 .xlsx 名
  title: ブックのタイトル帯（各シート 1 行目）
  meta:  補足（各シート 2 行目）
  sheets:
    - name: シート名
      columns:
        - { key: no, label: "No", width: 6, align: center }
        - { key: name, label: "項目名", width: 20 }
        - { key: phys, label: "物理名", width: 16, mono: true }
        - { key: req, label: "必須", width: 7, align: center, kind: required }
        - { key: note, label: "備考", width: 36, wrap: true }
      rows:
        - [1, "ユーザー ID", "userId", "○", "半角英数。前後空白はトリム"]

列メタ: `mono`=等幅 / `wrap`=折返し / `align`=寄せ / `kind: required`（○ を赤太字中央）/
        `kind: result`（合格=緑 / 不合格=赤 / 未=グレー、条件付き書式）。
"""
from __future__ import annotations

import pathlib
import re
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── デザイントークン（Word 版と共通。結果色のみ Excel 用に追加）──
JP = "BIZ UDPGothic"      # 本文
MONO = "BIZ UDGothic"     # 等幅（物理名・型・初期値・コード）
ACCENT = "1F5C99"         # ヘッダ塗り
ZEBRA = "F5F8FB"          # 偶数データ行
GRID = "C9D0D8"           # 罫線
INK = "20242C"            # 本文文字
MUTED = "606874"          # 補足
OK_TXT, OK_FILL = "1F8A5B", "E8F5EE"      # 合格
NG_TXT, NG_FILL = "B3403F", "FDECEC"      # 不合格・必須
TODO_TXT, TODO_FILL = "6B727B", "EEF1F4"  # 未

_THIN = Side(style="thin", color=GRID)
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)

TITLE_ROW = 1
META_ROW = 2
HEADER_ROW = 4   # この行で freeze & filter。データは 5 行目から。


# ─────────────────────────────────────────────── YAML サブセット・パーサ
# 対応する範囲（本生成器の原稿が使う構文のみ）:
#   - ブロックマッピング `key: value` / `key:`（直下にネスト）
#   - ブロックシーケンス `- ...`（スカラ / フローコレクション / `- key: ...` のマップ項目）
#   - フローシーケンス `[a, "b c", 3]` / フローマッピング `{ k: v, k2: "v2" }`（ネスト・クォート対応）
#   - スカラ: int / float / true|false / null|~ / "..."・'...'（クォート）/ 素の文字列
def parse_yaml(text: str):
    """YAML サブセット文字列を Python の dict/list/スカラへ変換する。"""
    raw = []
    for ln in text.splitlines():
        if not ln.strip() or _is_comment(ln):
            continue
        indent = len(ln) - len(ln.lstrip(" "))
        raw.append((indent, ln.strip()))
    value, _ = _parse_node(raw, 0, raw[0][0]) if raw else (None, 0)
    return value


def _is_comment(line: str) -> bool:
    s = line.lstrip()
    return s.startswith("#")


def _parse_node(lines, idx, indent):
    """`indent` 桁のブロック（マップ or シーケンス）を 1 つ読む。(value, next_idx) を返す。"""
    _, text = lines[idx]
    if text == "-" or text.startswith("- "):
        return _parse_seq(lines, idx, indent)
    return _parse_map(lines, idx, indent)


def _parse_seq(lines, idx, indent):
    seq = []
    while idx < len(lines):
        ind, text = lines[idx]
        if ind != indent or not (text == "-" or text.startswith("- ")):
            break
        item = text[1:].strip()
        if item == "":
            # 次行以降のより深いブロックが要素本体
            child = lines[idx + 1][0]
            val, idx = _parse_node(lines, idx + 1, child)
            seq.append(val)
        elif item[0] in "[{":
            seq.append(_parse_flow(item))
            idx += 1
        elif _looks_like_pair(item):
            # `- key: ...` 形式のマップ項目。先頭ペアはインライン、残りは indent+2 に揃う。
            val, idx = _parse_map(lines, idx, indent + 2, first=item)
            seq.append(val)
        else:
            seq.append(_parse_scalar(item))
            idx += 1
    return seq, idx


def _parse_map(lines, idx, indent, first=None):
    m = {}
    if first is not None:
        key, rest = _split_kv(first)
        idx += 1   # `- key: ...` の行を消費
        idx = _assign(lines, idx, indent, m, key, rest)
    while idx < len(lines):
        ind, text = lines[idx]
        if ind != indent or text.startswith("- "):
            break
        key, rest = _split_kv(text)
        idx += 1
        idx = _assign(lines, idx, indent, m, key, rest)
    return m, idx


def _assign(lines, idx, indent, m, key, rest):
    """`key: rest` の値を解決して `m[key]` に入れる。ネストブロックなら子を読み進める。"""
    if rest == "":
        if idx < len(lines) and lines[idx][0] > indent:
            child = lines[idx][0]
            val, idx = _parse_node(lines, idx, child)
        else:
            val = None
    elif rest[0] in "[{":
        val = _parse_flow(rest)
    else:
        val = _parse_scalar(rest)
    m[key] = val
    return idx


def _looks_like_pair(s: str) -> bool:
    return bool(re.match(r'^[^:\s"\'\[{][^:]*:(\s|$)', s))


def _split_kv(text: str):
    k, _, v = text.partition(":")
    return k.strip(), v.strip()


# ── フローコレクション（`[...]` / `{...}`）の再帰スキャナ ──
def _parse_flow(s: str):
    val, _ = _flow(s, 0)
    return val


def _skip_ws(s, i):
    while i < len(s) and s[i] in " \t":
        i += 1
    return i


def _flow(s, i):
    i = _skip_ws(s, i)
    if i >= len(s):
        return None, i
    if s[i] == "[":
        arr = []
        i = _skip_ws(s, i + 1)
        if s[i] == "]":
            return arr, i + 1
        while True:
            v, i = _flow(s, i)
            arr.append(v)
            i = _skip_ws(s, i)
            if i < len(s) and s[i] == ",":
                i = _skip_ws(s, i + 1)
                continue
            return arr, i + 1  # ']' を飛ばす
    if s[i] == "{":
        obj = {}
        i = _skip_ws(s, i + 1)
        if s[i] == "}":
            return obj, i + 1
        while True:
            key, i = _read_token(s, i, ":")
            i = _skip_ws(s, i)
            i += 1  # ':' を飛ばす
            v, i = _flow(s, i)
            obj[str(key).strip()] = v
            i = _skip_ws(s, i)
            if i < len(s) and s[i] == ",":
                i = _skip_ws(s, i + 1)
                continue
            return obj, i + 1  # '}' を飛ばす
    return _read_token(s, i, ",]}")


def _read_token(s, i, stops):
    """クォート文字列、または `stops` のいずれかまでの素トークンを 1 つ読み、型変換して返す。"""
    i = _skip_ws(s, i)
    if i < len(s) and s[i] in "\"'":
        quote = s[i]
        j = i + 1
        buf = []
        while j < len(s) and s[j] != quote:
            buf.append(s[j])
            j += 1
        return "".join(buf), j + 1
    j = i
    while j < len(s) and s[j] not in stops:
        j += 1
    return _parse_scalar(s[i:j].strip()), j


def _parse_scalar(token: str):
    t = token.strip()
    if len(t) >= 2 and t[0] in "\"'" and t[-1] == t[0]:
        return t[1:-1]
    low = t.lower()
    if low in ("true", "false"):
        return low == "true"
    if low in ("null", "~", ""):
        return None
    if re.fullmatch(r"[+-]?\d+", t):
        return int(t)
    if re.fullmatch(r"[+-]?\d*\.\d+", t):
        return float(t)
    return t


# ─────────────────────────────────────────────── ワークブック組み立て
def build_workbook(spec: dict, out_path):
    """パース済み `spec` から .xlsx を生成して `out_path` へ保存する。"""
    out_path = pathlib.Path(out_path)
    wb = Workbook()
    sheets = spec.get("sheets") or []
    if not sheets:
        raise ValueError("sheets が空です")
    for si, sh in enumerate(sheets):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = _safe_sheet_name(sh.get("name") or f"Sheet{si + 1}")
        _build_sheet(ws, sh, spec.get("title", ""), spec.get("meta", ""))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _safe_sheet_name(name: str) -> str:
    # Excel のシート名禁止文字を除去し 31 文字に丸める。
    return re.sub(r'[:\\/?*\[\]]', "_", str(name))[:31]


def _build_sheet(ws, sheet, book_title, book_meta):
    columns = sheet.get("columns") or []
    rows = sheet.get("rows") or []
    ncols = len(columns)
    last_col = get_column_letter(ncols)

    # 1〜2 行目: タイトル / メタ（結合セル）。
    ws.cell(TITLE_ROW, 1, sheet.get("title") or book_title)
    tcell = ws.cell(TITLE_ROW, 1)
    tcell.font = Font(name=JP, size=14, bold=True, color=INK)
    tcell.alignment = Alignment(vertical="center")
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")
    meta_text = sheet.get("meta") or book_meta
    if meta_text:
        ws.cell(META_ROW, 1, meta_text)
        mcell = ws.cell(META_ROW, 1)
        mcell.font = Font(name=JP, size=10, color=MUTED)
        mcell.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{META_ROW}:{last_col}{META_ROW}")

    # 4 行目: 列ヘッダ（ACCENT 塗り・白太字・中央）。
    for ci, col in enumerate(columns, start=1):
        c = ws.cell(HEADER_ROW, ci, col.get("label", col.get("key", "")))
        c.fill = HEADER_FILL
        c.font = Font(name=JP, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
        width = col.get("width")
        if width:
            ws.column_dimensions[get_column_letter(ci)].width = width

    # 5 行目〜: データ（ゼブラ＋細罫線）。
    first_data = HEADER_ROW + 1
    for ri, row in enumerate(rows):
        excel_row = first_data + ri
        zebra = (ri + 1) % 2 == 0   # データ 1 件目=奇数 → 偶数件目に zebra
        ws.row_dimensions[excel_row].height = max(18, ws.row_dimensions[excel_row].height or 0)
        for ci, col in enumerate(columns, start=1):
            val = row[ci - 1] if ci - 1 < len(row) else ""
            c = ws.cell(excel_row, ci, val)
            c.border = BORDER_ALL
            font_name = MONO if col.get("mono") else JP
            c.font = Font(name=font_name, size=10.5, color=INK)
            c.alignment = Alignment(
                vertical="center", wrap_text=bool(col.get("wrap")),
                horizontal=col.get("align") or "left")
            if zebra:
                c.fill = ZEBRA_FILL
            if col.get("kind") == "required" and str(val).strip() == "○":
                c.font = Font(name=JP, size=10.5, bold=True, color=NG_TXT)
                c.alignment = Alignment(horizontal="center", vertical="center")

    last_row = first_data + len(rows) - 1 if rows else HEADER_ROW

    # 結果列の条件付き書式（合格 / 不合格 / 未）。
    for ci, col in enumerate(columns, start=1):
        if col.get("kind") == "result" and rows:
            letter = get_column_letter(ci)
            rng = f"{letter}{first_data}:{letter}{last_row}"
            _add_result_rules(ws, rng)

    # 固定・フィルタ・印刷設定。
    ws.freeze_panes = ws.cell(row=first_data, column=1)
    if rows:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_row}"
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"
    ws.print_area = f"A1:{last_col}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.showGridLines = False


def _add_result_rules(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"合格"'],
        font=Font(name=JP, bold=True, color=OK_TXT), fill=PatternFill("solid", fgColor=OK_FILL)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"不合格"'],
        font=Font(name=JP, bold=True, color=NG_TXT), fill=PatternFill("solid", fgColor=NG_FILL)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"未"'],
        font=Font(name=JP, bold=True, color=TODO_TXT), fill=PatternFill("solid", fgColor=TODO_FILL)))


def render(yaml_path, out_path):
    """原稿 `yaml_path` を読み、`out_path` の .xlsx を生成する。戻り値は spec 辞書。"""
    yaml_path = pathlib.Path(yaml_path)
    spec = parse_yaml(yaml_path.read_text(encoding="utf-8"))
    build_workbook(spec, out_path)
    return spec


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: python md2xlsx.py <in.xlsx.yaml> <out.xlsx>", file=sys.stderr)
        raise SystemExit(2)
    render(sys.argv[1], sys.argv[2])
    print("wrote", sys.argv[2])
